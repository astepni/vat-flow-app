from datetime import datetime
from decimal import Decimal

import openpyxl
from django import forms
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView, TemplateView
from django.views.generic.edit import FormView
from invoices.models import Invoice
from openpyxl.utils import get_column_letter

from .models import Faktura
from .utils import FakturaCSVImporter, FakturaParser


class VATViewRegister(TemplateView):
    template_name = "vat_simulation/rejestr_vat.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        data_od = self.request.GET.get("data_od")
        data_do = self.request.GET.get("data_do")

        faktury = Faktura.objects.all()
        if data_od and data_do:
            faktury = faktury.filter(data_sprzedazy__range=[data_od, data_do])

        sprzedazowe = faktury.filter(typ="sprzedazowa")
        kosztowe = faktury.filter(typ="kosztowa")

        sprzedazowe_lista = []
        for f in sprzedazowe:
            dane = FakturaParser(f.plik_pdf.path).get_data()
            sprzedazowe_lista.append(
                {
                    **dane,
                    "plik_pdf": f.plik_pdf.url,
                    "pk": f.pk,
                }
            )

        kosztowe_lista = []
        for f in kosztowe:
            dane = FakturaParser(f.plik_pdf.path).get_data()
            kosztowe_lista.append(
                {
                    **dane,
                    "plik_pdf": f.plik_pdf.url,
                    "pk": f.pk,
                }
            )

        context["sprzedazowe_lista"] = sprzedazowe_lista
        context["kosztowe_lista"] = kosztowe_lista
        context["data_od"] = data_od
        context["data_do"] = data_do
        return context


class VatViewSimulation(LoginRequiredMixin, View):
    template_name = "vat_simulation/vat_simulation.html"

    def get(self, request):
        faktury = Faktura.objects.filter(user=request.user).order_by(
            "-data_wystawienia"
        )
        context = {
            "faktury": faktury,
            "suma_vat_nalezny": None,
            "suma_vat_naliczone": None,
            "saldo_vat": None,
            "okres_start": None,
            "okres_koniec": None,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        faktury = Faktura.objects.filter(user=request.user).order_by(
            "-data_wystawienia"
        )
        okres_start = request.POST.get("okres_start")
        okres_koniec = request.POST.get("okres_koniec")
        suma_vat_nalezny = None
        suma_vat_naliczone = None
        saldo_vat = None

        if okres_start and okres_koniec:
            data_start = datetime.strptime(okres_start, "%Y-%m-%d").date()
            data_koniec = datetime.strptime(okres_koniec, "%Y-%m-%d").date()

            vat_przychodowy = (
                Faktura.objects.filter(
                    user=request.user,
                    typ="sprzedazowa",
                    data_wystawienia__range=(data_start, data_koniec),
                ).aggregate(Sum("kwota_vat"))["kwota_vat__sum"]
                or 0
            )

            vat_kosztowy = (
                Faktura.objects.filter(
                    user=request.user,
                    typ="kosztowa",
                    data_wystawienia__range=(data_start, data_koniec),
                ).aggregate(Sum("kwota_vat"))["kwota_vat__sum"]
                or 0
            )

            suma_vat_nalezny = vat_przychodowy
            suma_vat_naliczone = vat_kosztowy
            saldo_vat = suma_vat_nalezny - suma_vat_naliczone

        context = {
            "faktury": faktury,
            "suma_vat_nalezny": suma_vat_nalezny,
            "suma_vat_naliczone": suma_vat_naliczone,
            "saldo_vat": saldo_vat,
            "okres_start": okres_start,
            "okres_koniec": okres_koniec,
        }
        return render(request, self.template_name, context)


class UploadCSVForm(forms.Form):
    csv_file = forms.FileField(label="Wybierz plik CSV")


class ImportInvoicesView(FormView):
    template_name = "vat_simulation/import_faktury.html"
    form_class = UploadCSVForm
    success_url = reverse_lazy("import_success")

    def form_valid(self, form):
        csv_file = form.cleaned_data["csv_file"]
        with open("temp.csv", "wb+") as temp_file:
            for chunk in csv_file.chunks():
                temp_file.write(chunk)
        importer = FakturaCSVImporter("temp.csv")
        importer.import_faktury()
        return super().form_valid(form)


class ApproveInvoiceView(View):
    def post(self, request, pk, *args, **kwargs):
        invoice = get_object_or_404(Invoice, pk=pk)
        dane = FakturaParser(invoice.pdf.path).get_data()
        raw_data_wystawienia = dane.get("data_wystawienia")
        raw_data_sprzedazy = dane.get("data_sprzedazy")

        def parse_any_date(value):
            for fmt in ("%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(value, fmt).date()
                except Exception:
                    continue
            return None

        def parse_decimal(value):
            if not value:
                return Decimal("0.00")
            value = value.replace(" ", "").replace(",", ".")
            try:
                return Decimal(value)
            except Exception:
                return Decimal("0.00")

        data_wystawienia = (
            parse_any_date(raw_data_wystawienia) if raw_data_wystawienia else None
        )
        data_sprzedazy = (
            parse_any_date(raw_data_sprzedazy) if raw_data_sprzedazy else None
        )

        if not data_wystawienia:
            return HttpResponse(
                "Brakuje daty wystawienia faktury do rejestru VAT albo format daty nieobsługiwany",
                status=400,
            )

        if Faktura.objects.filter(numer=dane.get("numer"), user=invoice.user).exists():
            return HttpResponse(
                "Taka faktura już istnieje w rejestrze VAT!", status=409
            )

        Faktura.objects.create(
            user=invoice.user,
            invoice=invoice,
            typ="sprzedazowa" if invoice.invoice_type == "income" else "kosztowa",
            numer=dane.get("numer"),
            data_wystawienia=data_wystawienia,
            data_sprzedazy=data_sprzedazy,
            kontrahent=dane.get("kontrahent"),
            adres_siedziba=dane.get("adres_siedziba"),
            wart_brutto_23=parse_decimal(dane.get("wart_brutto_23")),
            vat_23=parse_decimal(dane.get("podatek_vat_23")),
            netto_23=parse_decimal(dane.get("wart_netto_23")),
            wart_brutto_8=parse_decimal(dane.get("wart_brutto_8")),
            vat_8=parse_decimal(dane.get("podatek_vat_8")),
            netto_8=parse_decimal(dane.get("wart_netto_8")),
            suma_netto=parse_decimal(dane.get("suma_wart_netto")),
            suma_vat=parse_decimal(dane.get("suma_podatek_vat")),
            suma_brutto=parse_decimal(dane.get("suma_wart_brutto")),
            rodzaj="sprzedazowa" if invoice.invoice_type == "income" else "kosztowa",
            plik_pdf=invoice.pdf,
        )
        return redirect("vat_simulation:rejestr_vat")


class InvoiceListView(LoginRequiredMixin, ListView):
    model = Faktura
    template_name = "vat_simulation/invoice_list.html"
    context_object_name = "faktury"

    def get_queryset(self):
        return Faktura.objects.filter(user=self.request.user).order_by(
            "-data_wystawienia"
        )


class DeleteInvoiceView(View):
    def post(self, request, pk, *args, **kwargs):
        faktura = get_object_or_404(Faktura, pk=pk)
        faktura.delete()
        return redirect("vat_simulation:rejestr_vat")


def convert_number(val):
    if isinstance(val, (float, int)):
        return val
    if not val:
        return 0
    return float(str(val).replace(" ", "").replace(",", "."))


class ExportExcelView(View):
    def get(self, request):
        data_od = request.GET.get("data_od")
        data_do = request.GET.get("data_do")

        faktury = Faktura.objects.all()
        if data_od and data_do and data_od != "None" and data_do != "None":
            faktury = faktury.filter(data_sprzedazy__range=[data_od, data_do])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rejestr VAT"

        columns = [
            "Nr faktury",
            "Data wystawienia",
            "Data sprzedaży",
            "Kontrahent",
            "Adres",
            "Brutto 23%",
            "VAT 23%",
            "Netto 23%",
            "Brutto 8%",
            "VAT 8%",
            "Netto 8%",
            "Suma netto",
            "Suma VAT",
            "Suma brutto",
        ]
        ws.append(columns)

        for f in faktury:
            row = [
                f.numer,
                f.data_wystawienia.strftime("%d.%m.%Y") if f.data_wystawienia else "",
                f.data_sprzedazy.strftime("%d.%m.%Y") if f.data_sprzedazy else "",
                f.kontrahent,
                f.adres_siedziba,
                convert_number(f.wart_brutto_23),
                convert_number(f.vat_23),
                convert_number(f.netto_23),
                convert_number(f.wart_brutto_8),
                convert_number(f.vat_8),
                convert_number(f.netto_8),
                convert_number(f.suma_netto),
                convert_number(f.suma_vat),
                convert_number(f.suma_brutto),
            ]
            ws.append(row)

        for col_num, column_title in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=Rejestr_VAT.xlsx"
        wb.save(response)
        return response
